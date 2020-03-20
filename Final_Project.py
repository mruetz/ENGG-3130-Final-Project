# ENGG*3130 - Modeling Complex Systems
## Final Project

import io
import networkx as nx
import numpy as np
import json
import string
import matplotlib.pyplot as plt
from itertools import combinations 

# Import MapQuest API Libraries from https://github.com/MapQuest/directions-api-python-client somehow
from RouteOptions import RouteOptions
from RouteService import RouteService

# Import Excel libraries
from openpyxl import Workbook
from openpyxl import load_workbook

#%matplotlib inline

data_file = "./data/1_02.xlsm"

wb = load_workbook(data_file, read_only=False, keep_vba=True)
ws = wb.active
deliveries = {}
load_weights = []

options = RouteOptions()
service = RouteService('dAGkCaPVqA3OcC5ws2Lfv8wsR9ro45oe')
# our key is "dAGkCaPVqA3OcC5ws2Lfv8wsR9ro45oe"

for row in range(1, len(ws['A'])+1):
    if ('delivery' in str(ws['J'+str(row)].value)):
        if ws['A'+str(row)].value in deliveries.keys():
            deliveries[ws['A'+str(row)].value].append(ws['R'+str(row)].value)
        else:
            deliveries[ws['A'+str(row)].value] = [ws['R'+str(row)].value]
        load_weights.append(ws['AK'+str(row)].value)

one_truck_locations = deliveries[1646] #Enter number for truck to get locations for
for loc in one_truck_locations:
    print loc



def get_weighted_dist_matrix(locations, loads, factor_weight):

    route_matrix = service.routeMatrix(locations=locations, allToAll=True)

    print(route_matrix)

    dist_matrix = route_matrix['distance']
    weighted_dist_matrix = dist_matrix

    i = 0;
    for load_weight in loads:
        weighted_dist_matrix[i] = dist_matrix[i]*(1/load_weight)*factor_weight
        ++i
    return weighted_dist_matrix


def get_weighted_time_matrix(locations, loads, factor_weight):

    route_matrix = service.routeMatrix(locations=locations, allToAll=true)

    time_matrix = route_matrix['time']
    weighted_time_matrix = time_matrix

    i = 0;
    for load_weight in loads:
        weighted_time_matrix[i] = time_matrix[i]*(1/load_weight)* factor_weight
        ++i
    return weighted_time_matrix

"""
def calc_edge(addr_1, addr_2, factor=None):
	#Calculate the edge weight based on certain factors

	service = RouteService('YOUR_MAPQUEST_DEVELOPER_API_KEY')
	location_list=[addr_1, addr_2]

	if factor is None or factor is 'dist':
		# MapQuest distance from addr_1 to addr_2
		routeMatrix = service.routeMatrix(locations=location_list, allToAll=True)
		return routeMatrix['distance'][0][1]

	elif factor is 'time':
		return 0
		# Return MapQuest time

	#elif factor is 'dist/time':
		# MapQuest time divided by distance

	#elif factor is 'fuel':
		# Mapquest distance from l1 to l2 
		# Taking weight into account for fuel efficiency
		# Use the weight_point_for_load() method
	"""
  
def create_individual_driver_graphs(nodes):
	G = nx.DiGraph()
	drive_times = {}
	for driver in deliveries.keys():
		drive_times.clear()
		nodes = deliveries[driver]
		G.add_nodes_from(nodes)
		all_combs = combinations(nodes, 2) 
		for i in len(list(all_combs)):
			drive_times.add(list(all_combs)[i], calc_edge(list(all_combs)[i][0],list(all_combs)[i][1]), 'dist')
			G.add_edges_from(drive_times)
	
	# nx.shortest_path() investigation


weighted_dist_matrix = get_weighted_dist_matrix(one_truck_locations, load_weights, 0.5)

"Original Dist Matrix:"
route_matrix = service.routeMatrix(locations=one_truck_locations, allToAll=true)
print route_matrix
#dist_matrix = route_matrix['distance']
#print dist_matrix

print "Weighted Dist Matrix:"
print weighted_dist_matrix
